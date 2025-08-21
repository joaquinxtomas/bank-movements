import glob
import re
import openpyxl
from openpyxl.styles import Border, Side
from datetime import datetime
import locale
import pdfplumber

#border style if new row is added
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

def obt_datapdf(page):
    # extract all words with their coordinates from the PDF
    words = page.extract_words()

    # use coordinates to group words into lines in the PDF
    lines = {}
    for word in words:
        y_coord = int(word['top'])
        if y_coord not in lines:
            lines[y_coord] = []
        lines[y_coord].append(word)
    
    final_rows = []
    # iterate over lines vertically
    for y_coord in sorted(lines.keys()):
        # sort words by horizontal coordinate to preserve the PDF order
        line_words = sorted(lines[y_coord], key = lambda x:x['x0'])
        
        fecha = ""
        concepto = []
        importe = ""
        
        # assign values to variables based on horizontal coordinates
        # 20 to 70: date
        # 80 to 300: concept (largest column)
        # 310 to 380: amount
        for word in line_words:
            x_coord = word['x0']
            if 25 <= x_coord <= 70:
                fecha = word['text']
            elif 80 <= x_coord <= 300:
                concepto.append(word['text'])
            elif 310 <= x_coord <= 380:
                importe = word['text']
                
        # exclude lines that are not data rows
        if fecha and importe:
            if(concepto != "Concepto" and importe != "Importe" and fecha != "Fecha"):
                final_rows.append([fecha, " ".join(concepto), importe])
    return final_rows

# get all sheet names from the .xlsx file and return as a list
def obt_sheets(arch):
    try:
        book = openpyxl.load_workbook(arch)
        return book.sheetnames
    except FileNotFoundError:
        return f"Error, el archivo '{arch}' no fue encontrado"
    
# return the month name from a string date (dd/mm/yyyy)
def obt_month(str_fecha):
    locale.setlocale(locale.LC_TIME, "Spanish_Spain")
    fecha = datetime.strptime(str_fecha, "%d/%m/%Y")
    mes = fecha.strftime("%B").upper()
    return mes
    
# keys represent the target column for each type of data
# values are keywords that may appear in the PDF
column_map = {
    2:["DEPOSITO EFECTIVO", "DEPOSITO POR CAJA"],
    3:["CRED.TRF", "CR.TRAN", "BIP CR TR", "CR.DEBIN"],
    4:["DEPOSITO CHEQUE"],
    5:["COMISION"],
    6:["INTERESES"],
    7:["IMPUESTO SELLOS"],
    8:["IMPUESTO DEBITO"],
    9:["IMPUESTO CREDITO"],
    10:["SIRCREB","IMPUESTO I.BRUTOS - PERCEPCION"],
    11:["INGRESOS BRUTOS - PERCEPCION"],
    12:["P.SERV", "FEDERACION PATR", "CIRCULO CERRADO", "COMPRA TARJETA",
        "METROGAS"],
    13:["IIBB"],
    14:["SINDICATO"],
    15:["SOUTO"],
    16:["PAGO CHEQUE", "CHEQUE DE CAMARA"],
    17:["PAGO AFIP", "VEP"],
    18:["BIP DB.TR","BIP DB TR","DB.DEBIN","DEB.TRF", "PAGO LIQUIDACION VISA"],
}

txt_file = glob.glob("*.txt")
excel_file = glob.glob("*.xlsx")
pdf_files = glob.glob("*.pdf")

lines = []

if(pdf_files and txt_file):
    print("No puede haber un archivo .TXT y uno .PDF al mismo tiempo en la carpeta.")
elif (pdf_files):
    file = pdf_files[0]
    # extract lines from the PDF using obt_datapdf function and add to lines list
    with pdfplumber.open(file) as pdf:
        num_pages = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages):
            if page_num == num_pages - 1:
                continue
            lines_on_page = obt_datapdf(page)
            lines.extend(lines_on_page)
elif (txt_file):
    if(len(txt_file) == 1):
     filename = txt_file[0]
     with open(filename, "r", encoding="utf-8") as f:
         start = False
         # search for the header line "FECHA,CONCEPTO,IMPORTE"
         for x in f:
             x = x.strip()
             if not start:
                 if x.startswith("FECHA,CONCEPTO,IMPORTE"):
                     start = True
                     headers = x.split(",")
                     lines.append(headers)
                     # remove the first line since it's the header
                     lines.pop(0)
             else:
                 # if the line is not empty, split by commas to create parts list
                 if x != "":
                     parts = [p.strip() for p in x.split(",")]
                    
                    # some entries have an extra field (extra commas), remove the third element
                     if(len(parts) > 5):
                         parts.pop(2)
                    
                    # validate the date format
                     if(len(parts) > 0):
                         fecha = parts[0]
                         date_pattern = r"^\d{2}/\d{2}/\d{4}$"
                         match = re.match(date_pattern, fecha)
                         if(match):
                             lines.append(parts)            
else:
     print("Hay más de un archivo .txt en la carpeta, debe haber solo uno.")


if(len(excel_file) == 1) and lines:
    # load the excel file
    excel_filename = excel_file[0]
    wb = openpyxl.load_workbook(excel_filename)
    sheet_names = obt_sheets(excel_filename)
    
    # the second line defines the date, use it to select the correct month sheet
    sheet_date = lines[1][0]
    month = obt_month(sheet_date)
    for name in sheet_names:
        if(month == name):    
            worksheet = wb[month]
    
    start_index = None
    last_index = None
    # find cells containing "SALDO ANTERIOR" and "TOTALES" to write between them
    for row in worksheet.iter_rows(min_row=1, max_col=2):
        cell = row[1]
        if cell.value and "SALDO ANTERIOR" in str(cell.value).upper():
            start_index = cell.row + 1
        if cell.value and "TOTALES" in str(cell.value):
            last_index = cell.row

    row_to_fill = start_index
    
    for line in lines:
        concepto = line[1].upper()
        col_excel = None
        
        # check if any keyword is in the dictionary to determine column
        for col, keywords in column_map.items():
            for kw in keywords:
                if kw and kw in concepto:
                    col_excel = col
                    break
            if(col_excel):
                break
        
        importe = abs(float(line[2]))
        fecha = line[0]
        
        #if the next row exceeds the last row, insert a new row 
        if(row_to_fill >= last_index):
            worksheet.insert_rows(last_index)
            last_index += 1
        
        # assign values to the corresponding columns
        if(obt_month(fecha) == month):
            if(col_excel is not None):
                worksheet.cell(row=row_to_fill, column=1, value=fecha)
                worksheet.cell(row=row_to_fill, column=2, value=concepto)
                worksheet.cell(row=row_to_fill,column=col_excel+1, value=importe)
            else:
                worksheet.cell(row=row_to_fill, column=3, value="")

            # apply borders to columns 1 to 20
            for col1 in range(1, 21):
                cell = worksheet.cell(row=row_to_fill, column=col1)
                cell.border = thin_border

            # move to the next row for the next entry
            row_to_fill += 1
    
    # save the excel file 
    wb.save(excel_filename)

